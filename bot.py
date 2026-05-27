import logging
import openpyxl
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler

# Configuracion del log para ver errores en consola
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Token del bot
TOKEN = "8091309523:AAEEy4I51oquoUl7gExYUxEkfDX208Fw1r4"

# Ruta del archivo Excel
EXCEL_PATH = "Base de datos.xlsx"

# Estados de la maquina de estados
NOMBRE, ID_EMPLEADO, DESCRIPCION, TIPO_PROBLEMA = range(4)

# Prioridad segun tipo de problema
def obtener_prioridad(tipo):
    prioridades = {
        "Hardware": "Alta",
        "Red": "Alta",
        "Software": "Media",
        "Otro": "Baja"
    }
    return prioridades.get(tipo, "Baja")

# Verificar si el empleado existe en el Excel
def verificar_empleado(id_empleado):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Empleados"]
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if str(fila[0]) == str(id_empleado):
            return fila[1]  # Devuelve el nombre
    return None

# Registrar ticket en el Excel
def registrar_ticket(id_empleado, tipo, descripcion, prioridad):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Tickets"]
    # Generar ID de ticket
    ultimo = ws.max_row
    id_ticket = f"TKT{ultimo:03d}"
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.append([id_ticket, id_empleado, tipo, descripcion, prioridad, "Abierto", fecha])
    wb.save(EXCEL_PATH)
    return id_ticket

# /start - Inicia el proceso
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Bienvenido al sistema de Soporte Técnico.\n\n"
        "Voy a ayudarte a registrar tu ticket de soporte.\n\n"
        "¿Cuál es tu nombre?"
    )
    return NOMBRE

# Recibe el nombre
async def recibir_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()
    if not texto.replace(" ", "").isalpha():
        await update.message.reply_text("⚠️ El nombre solo puede contener letras. Intentá de nuevo:")
        return NOMBRE
    context.user_data["nombre"] = texto
    await update.message.reply_text(
        f"Hola {texto} 👋\n\n"
        "Por favor ingresá tu número de empleado (ej: EMP001):"
    )
    return ID_EMPLEADO

# Recibe y valida el ID de empleado
async def recibir_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    id_emp = update.message.text.strip().upper()
    nombre = verificar_empleado(id_emp)
    if nombre is None:
        await update.message.reply_text(
            "❌ No encontré ese número de empleado en el sistema.\n\n"
            "Verificá el número e intentá de nuevo, o contactá a RRHH para registrarte.\n\n"
            "Ingresá tu número de empleado:"
        )
        return ID_EMPLEADO
    context.user_data["id_empleado"] = id_emp
    await update.message.reply_text(
        f"✅ Empleado verificado: {nombre}\n\n"
        "Describí brevemente el problema que tenés:"
    )
    return DESCRIPCION

# Recibe la descripcion del problema
async def recibir_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["descripcion"] = update.message.text.strip()
    teclado = [["Hardware", "Software"], ["Red", "Otro"]]
    markup = ReplyKeyboardMarkup(teclado, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(
        "¿Qué tipo de problema es?",
        reply_markup=markup
    )
    return TIPO_PROBLEMA

# Recibe el tipo de problema y genera el ticket
async def recibir_tipo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tipo = update.message.text.strip()
    tipos_validos = ["Hardware", "Software", "Red", "Otro"]
    if tipo not in tipos_validos:
        teclado = [["Hardware", "Software"], ["Red", "Otro"]]
        markup = ReplyKeyboardMarkup(teclado, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text(
            "⚠️ Por favor seleccioná una opción del teclado:",
            reply_markup=markup
        )
        return TIPO_PROBLEMA
    prioridad = obtener_prioridad(tipo)
    id_ticket = registrar_ticket(
        context.user_data["id_empleado"],
        tipo,
        context.user_data["descripcion"],
        prioridad
    )
    await update.message.reply_text(
        f"✅ Ticket registrado exitosamente!\n\n"
        f"📋 Número de ticket: {id_ticket}\n"
        f"🔧 Tipo: {tipo}\n"
        f"⚡ Prioridad: {prioridad}\n\n"
        f"Un técnico se comunicará con vos a la brevedad.\n"
        f"Gracias por usar el sistema de soporte. 🙌",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END

# Comando cancelar
async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❌ Operación cancelada. Escribí /start para comenzar de nuevo.",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END

# Inicio de la aplicacion
def main():
    app = Application.builder().token(TOKEN).build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            NOMBRE: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_nombre)],
            ID_EMPLEADO: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_id)],
            DESCRIPCION: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_descripcion)],
            TIPO_PROBLEMA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_tipo)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)]
    )
    app.add_handler(conv_handler)
    print("Bot iniciado. Presiona Ctrl+C para detenerlo.")
    app.run_polling()

if __name__ == "__main__":
    main()